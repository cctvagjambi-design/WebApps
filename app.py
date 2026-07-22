import os
from datetime import date, datetime
from functools import wraps
from io import BytesIO

from flask import Flask, flash, g, redirect, render_template, request, Response, session, url_for
from flask_sqlalchemy import SQLAlchemy
from openpyxl import Workbook, load_workbook
from sqlalchemy import inspect, or_, text
from werkzeug.security import check_password_hash, generate_password_hash

app = Flask(__name__)
os.makedirs(app.instance_path, exist_ok=True)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "change-this-secret")
database_path = os.path.join(app.instance_path, "app.db")
app.config["SQLALCHEMY_DATABASE_URI"] = f"sqlite:///{database_path.replace('\\', '/')}"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)

STATUS_CHOICES = [
    "terdaftar",
    "mencetak",
    "menyiapkan barang",
    "mengirim",
    "mengupload",
]


class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    is_root = db.Column(db.Boolean, default=False, nullable=False)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)


class DeliverySlip(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    slip_number = db.Column(db.String(120), unique=True, nullable=False)
    customer_name = db.Column(db.String(120), nullable=False, default="")
    equipment = db.Column(db.String(120), nullable=False, default="")
    delivery_date = db.Column(db.Date, nullable=False, default=date.today)
    description = db.Column(db.Text, nullable=True)
    status = db.Column(db.String(50), nullable=False, default="terdaftar")
    created_by_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    created_by = db.relationship("User", backref="delivery_slips")
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow, onupdate=datetime.utcnow)


class Equipment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    equipment_number = db.Column(db.String(120), unique=True, nullable=False)
    serial_number = db.Column(db.String(120), nullable=True)
    model = db.Column(db.String(120), nullable=True)
    customer_name = db.Column(db.String(200), nullable=True)
    address = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)


def ensure_delivery_slip_columns():
    inspector = inspect(db.engine)
    if "delivery_slip" not in inspector.get_table_names():
        return

    existing_columns = {column["name"] for column in inspector.get_columns("delivery_slip")}
    if "customer_name" not in existing_columns:
        db.session.execute(text("ALTER TABLE delivery_slip ADD COLUMN customer_name VARCHAR(120) NOT NULL DEFAULT ''"))
    if "equipment" not in existing_columns:
        db.session.execute(text("ALTER TABLE delivery_slip ADD COLUMN equipment VARCHAR(120) NOT NULL DEFAULT ''"))
    if "delivery_date" not in existing_columns:
        db.session.execute(text("ALTER TABLE delivery_slip ADD COLUMN delivery_date DATE"))
    db.session.commit()


def init_db():
    db.create_all()
    ensure_delivery_slip_columns()
    if not User.query.filter_by(username="root").first():
        root_password = os.environ.get("ROOT_PASSWORD", "root")
        root = User(
            username="root",
            password_hash=generate_password_hash(root_password),
            is_root=True,
        )
        db.session.add(root)
        db.session.commit()
        app.logger.info("Created default root user 'root'. Set ROOT_PASSWORD to change the password.")


with app.app_context():
    init_db()


@app.before_request
def load_logged_in_user():
    user_id = session.get("user_id")
    g.user = User.query.get(user_id) if user_id else None


def login_required(view):
    @wraps(view)
    def wrapped_view(*args, **kwargs):
        if g.user is None:
            flash("You need to login first.", "warning")
            return redirect(url_for("login"))
        return view(*args, **kwargs)

    return wrapped_view


def root_required(view):
    @wraps(view)
    @login_required
    def wrapped_view(*args, **kwargs):
        if not g.user.is_root:
            flash("Access denied. Root user required.", "danger")
            return redirect(url_for("index"))
        return view(*args, **kwargs)

    return wrapped_view


def get_next_status(current_status):
    try:
        current_index = STATUS_CHOICES.index(current_status)
    except ValueError:
        return None
    if current_index + 1 < len(STATUS_CHOICES):
        return STATUS_CHOICES[current_index + 1]
    return None


def can_manage_delivery_slip():
    return g.user is not None


def normalize_cell_value(value):
    if value is None:
        return ""
    return str(value).strip()


def find_header_index(headers, *candidate_names):
    normalized_candidates = {name.lower(): name for name in candidate_names}
    for index, header in enumerate(headers):
        if header is None:
            continue
        header_text = normalize_cell_value(header).lower()
        if header_text in normalized_candidates:
            return index
    return None


def load_equipment_workbook(uploaded_file=None):
    if uploaded_file and getattr(uploaded_file, "filename", ""):
        return load_workbook(uploaded_file, read_only=True, data_only=True), uploaded_file.filename

    default_path = os.environ.get("DEFAULT_EQUIPMENT_IMPORT_PATH", r"C:\Users\kusyadi.ASTRAGRAPHIA\Dev\mif.xlsx")
    if os.path.exists(default_path):
        return load_workbook(default_path, read_only=True, data_only=True), default_path

    raise FileNotFoundError("No Excel file provided and the default import file was not found.")


@app.route("/")
@login_required
def index():
    slips = DeliverySlip.query.order_by(DeliverySlip.created_at.desc()).all()
    return render_template("index.html", slips=slips, status_choices=STATUS_CHOICES)


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"].strip()
        password = request.form["password"]
        user = User.query.filter_by(username=username).first()
        if user is None or not user.check_password(password):
            flash("Invalid username or password.", "danger")
            return render_template("login.html")
        session.clear()
        session["user_id"] = user.id
        flash(f"Welcome, {user.username}!", "success")
        return redirect(url_for("index"))
    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    session.clear()
    flash("You have been logged out.", "info")
    return redirect(url_for("login"))


@app.route("/users/register", methods=["GET", "POST"])
@root_required
def register_user():
    if request.method == "POST":
        username = request.form["username"].strip()
        password = request.form["password"]
        confirm_password = request.form["confirm_password"]
        if not username or not password:
            flash("Username and password are required.", "danger")
        elif password != confirm_password:
            flash("Passwords do not match.", "danger")
        elif User.query.filter_by(username=username).first():
            flash("Username already exists.", "danger")
        else:
            new_user = User(
                username=username,
                password_hash=generate_password_hash(password),
                is_root=False,
            )
            db.session.add(new_user)
            db.session.commit()
            flash(f"User '{username}' successfully registered.", "success")
            return redirect(url_for("index"))
    return render_template("register_user.html")


@app.route("/users")
@root_required
def list_users():
    users = User.query.order_by(User.username.asc()).all()
    return render_template("users.html", users=users)


@app.route("/change-password", methods=["GET", "POST"])
@login_required
def change_password():
    target_user = g.user
    if request.method == "POST":
        current_password = request.form.get("current_password", "")
        new_password = request.form.get("new_password", "")
        confirm_password = request.form.get("confirm_password", "")

        if not new_password or not confirm_password:
            flash("New password and confirmation are required.", "danger")
        elif new_password != confirm_password:
            flash("Passwords do not match.", "danger")
        else:
            can_change = True
            if not g.user.is_root:
                if not current_password or not target_user.check_password(current_password):
                    flash("Current password is incorrect.", "danger")
                    can_change = False
            elif current_password and not target_user.check_password(current_password):
                flash("Current password is incorrect.", "danger")
                can_change = False

            if can_change:
                target_user.password_hash = generate_password_hash(new_password)
                db.session.commit()
                flash("Password updated successfully.", "success")
                return redirect(url_for("index"))

    return render_template("change_password.html", target_user=target_user, is_self=True)


@app.route("/users/<int:user_id>/change-password", methods=["GET", "POST"])
@root_required
def change_user_password(user_id):
    target_user = User.query.get_or_404(user_id)
    if request.method == "POST":
        new_password = request.form.get("new_password", "")
        confirm_password = request.form.get("confirm_password", "")

        if not new_password or not confirm_password:
            flash("New password and confirmation are required.", "danger")
        elif new_password != confirm_password:
            flash("Passwords do not match.", "danger")
        else:
            target_user.password_hash = generate_password_hash(new_password)
            db.session.commit()
            flash(f"Password updated for user '{target_user.username}'.", "success")
            return redirect(url_for("list_users"))

    return render_template("change_password.html", target_user=target_user, is_self=False)


@app.route("/users/<int:user_id>/delete", methods=["POST"])
@root_required
def delete_user(user_id):
    target_user = User.query.get_or_404(user_id)
    if target_user.id == g.user.id:
        flash("You cannot delete your own account.", "danger")
        return redirect(url_for("list_users"))

    db.session.delete(target_user)
    db.session.commit()
    flash(f"User '{target_user.username}' deleted successfully.", "success")
    return redirect(url_for("list_users"))


@app.route("/delivery/export")
@login_required
def export_delivery_excel():
    slips = DeliverySlip.query.order_by(DeliverySlip.created_at.desc()).all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Delivery Slips"
    headers = ["Slip Number", "Customer Name", "Equipment", "Delivery Date", "Status", "Created By", "Updated At"]
    sheet.append(headers)

    for slip in slips:
        sheet.append(
            [
                slip.slip_number,
                slip.customer_name,
                slip.equipment,
                slip.delivery_date.strftime("%Y-%m-%d") if slip.delivery_date else "",
                slip.status,
                slip.created_by.username if slip.created_by else "",
                slip.updated_at.strftime("%Y-%m-%d %H:%M") if slip.updated_at else "",
            ]
        )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=delivery_slips.xlsx"},
    )


@app.route("/delivery/register", methods=["GET", "POST"])
@login_required
def register_delivery():
    if not can_manage_delivery_slip():
        flash("You need to login first.", "warning")
        return redirect(url_for("login"))

    if request.method == "POST":
        slip_number = request.form["slip_number"].strip()
        customer_name = request.form.get("customer_name", "").strip()
        equipment = request.form.get("equipment", "").strip()
        delivery_date = request.form.get("delivery_date", "").strip()
        if not slip_number:
            flash("Delivery slip number is required.", "danger")
        elif not customer_name:
            flash("Customer name is required.", "danger")
        elif not equipment:
            flash("Equipment is required.", "danger")
        elif not delivery_date:
            flash("Delivery date is required.", "danger")
        elif DeliverySlip.query.filter_by(slip_number=slip_number).first():
            flash("This delivery slip number is already registered.", "danger")
        else:
            try:
                parsed_date = datetime.strptime(delivery_date, "%Y-%m-%d").date()
            except ValueError:
                flash("Delivery date must be in YYYY-MM-DD format.", "danger")
            else:
                slip = DeliverySlip(
                    slip_number=slip_number,
                    customer_name=customer_name,
                    equipment=equipment,
                    delivery_date=parsed_date,
                    status="terdaftar",
                    created_by=g.user,
                )
                db.session.add(slip)
                db.session.commit()
                flash("Delivery slip successfully registered.", "success")
                return redirect(url_for("index"))
    return render_template("register_delivery.html")


@app.route("/delivery/<int:slip_id>/update", methods=["GET", "POST"])
@login_required
def update_delivery(slip_id):
    if not can_manage_delivery_slip():
        flash("You need to login first.", "warning")
        return redirect(url_for("login"))

    slip = DeliverySlip.query.get_or_404(slip_id)

    next_status = get_next_status(slip.status)
    if request.method == "POST":
        if not next_status:
            flash("This delivery slip is already in the final status.", "info")
        else:
            slip.status = next_status
            db.session.commit()
            flash(f"Status updated to '{next_status}'.", "success")
            return redirect(url_for("index"))
    return render_template("update_delivery.html", slip=slip, next_status=next_status)


@app.route("/tools/equipment")
@login_required
def list_equipment():
    query = request.args.get("q", "").strip()
    if query:
        filters = [
            Equipment.equipment_number.contains(query),
            Equipment.serial_number.contains(query),
            Equipment.model.contains(query),
            Equipment.customer_name.contains(query),
            Equipment.address.contains(query),
        ]
        equipment = Equipment.query.filter(or_(*filters)).order_by(Equipment.created_at.desc(), Equipment.equipment_number.asc()).all()
    else:
        equipment = Equipment.query.order_by(Equipment.created_at.desc(), Equipment.equipment_number.asc()).all()
    return render_template("equipment_list.html", equipment=equipment, query=query)


@app.route("/tools/import-equipment", methods=["GET", "POST"])
@login_required
def import_equipment():
    imported_count = None
    if request.method == "POST":
        uploaded_file = request.files.get("file")
        if not uploaded_file or uploaded_file.filename == "":
            flash("No file was selected, so the default import file will be used if it exists.", "info")

        try:
            workbook, source_name = load_equipment_workbook(uploaded_file)
            sheet = workbook.active
            imported_count = 0
            total_rows = max(1, sheet.max_row - 1)

            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [normalize_cell_value(value) for value in header_row]
            equipment_idx = find_header_index(headers, "equipment")
            serial_idx = find_header_index(headers, "serial no.", "serial number")
            model_idx = find_header_index(headers, "material description", "model no.", "model")
            customer_idx = find_header_index(headers, "customer", "list name", "customer name", "name pelanggan")
            address_idx = find_header_index(headers, "street", "city", "address", "alamat")

            fallback_indices = {
                "equipment": 3,
                "serial": 4,
                "model": 11,
                "customer": 14,
                "address": 15,
            }

            def get_value(row_values, index, fallback):
                if index is not None and index < len(row_values):
                    return normalize_cell_value(row_values[index])
                if fallback is not None and fallback < len(row_values):
                    return normalize_cell_value(row_values[fallback])
                return ""

            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_values = [normalize_cell_value(value) for value in row]
                if not any(row_values):
                    continue

                equipment_number = get_value(row_values, equipment_idx, fallback_indices["equipment"])
                serial_number = get_value(row_values, serial_idx, fallback_indices["serial"])
                model = get_value(row_values, model_idx, fallback_indices["model"])
                customer_name = get_value(row_values, customer_idx, fallback_indices["customer"])
                address = get_value(row_values, address_idx, fallback_indices["address"])

                if not equipment_number:
                    continue

                existing = Equipment.query.filter_by(equipment_number=equipment_number).first()
                if existing:
                    existing.serial_number = serial_number or existing.serial_number
                    existing.model = model or existing.model
                    existing.customer_name = customer_name or existing.customer_name
                    existing.address = address or existing.address
                else:
                    equipment = Equipment(
                        equipment_number=equipment_number,
                        serial_number=serial_number or None,
                        model=model or None,
                        customer_name=customer_name or None,
                        address=address or None,
                    )
                    db.session.add(equipment)
                imported_count += 1

            db.session.commit()
            flash(f"Imported {imported_count} equipment record(s) from {source_name}.", "success")
        except Exception as exc:
            flash(f"Unable to import file: {exc}", "danger")

    return render_template("import_equipment.html", imported_count=imported_count, total_rows=total_rows if 'total_rows' in locals() else None)


if __name__ == "__main__":
    host = os.environ.get("HOST", "0.0.0.0")
    port = int(os.environ.get("PORT", "5000"))
    debug = os.environ.get("FLASK_DEBUG", "False").lower() in {"1", "true", "yes", "on"}
    app.run(host=host, port=port, debug=debug)
