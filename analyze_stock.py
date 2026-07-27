from openpyxl import load_workbook

wb = load_workbook('C:\\Users\\kusyadi.ASTRAGRAPHIA\\Dev\\stock.xlsx', read_only=True)
sheet = wb.active

print('Headers:')
for i, cell in enumerate(sheet[1], 1):
    print(f'Column {i}: {cell.value}')

print('\nFirst 5 data rows:')
for row in sheet.iter_rows(min_row=2, max_row=6, values_only=True):
    print(row)
